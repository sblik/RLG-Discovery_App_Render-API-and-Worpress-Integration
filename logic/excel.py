"""
Excel spreadsheet generation for discovery documents.

Contains: Functions to create formatted Excel discovery indexes.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


def build_discovery_xlsx(
    df: pd.DataFrame,
    *,
    party: str = "Client",
    title_text: str = "CLIENT NAME - DOCUMENTS",
    date_col_name: str = "Date Produced",
    name_col_name: str = "Document Name/Title",
    cat_col_name: str = "Category",
    bates_col_name: str = "Bates Range",
) -> bytes:
    """
    Build a formatted Excel discovery index spreadsheet.

    Args:
        df: DataFrame with document information
        party: Party name ("Client" or "OP") for color coding
        title_text: Title text for the spreadsheet header
        date_col_name: Name of the date column
        name_col_name: Name of the document name column
        cat_col_name: Name of the category column
        bates_col_name: Name of the Bates range column

    Returns:
        Excel file bytes
    """
    logger.info("build_discovery_xlsx: %d document(s), party=%r, title=%r",
                len(df), party, title_text)

    PARTY_COLORS = {
        "Client": "FFEFFFF2",  # Light green
        "OP": "FFEDF7FF",  # Light blue
    }
    category_fill = PatternFill("solid", fgColor=PARTY_COLORS.get(party, "FFEFFFF2"))

    header_fill = PatternFill("solid", fgColor="FF1F4E79")
    header_font = Font(bold=True, color="FFFFFFFF")
    normal_font = Font(bold=False, color="FF000000")
    title_font = Font(bold=True, size=16)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="FFBBBBBB")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Client" if party == "Client" else "OP"

    ws.merge_cells("A1:C1")
    ws["A1"] = title_text
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    headers = [date_col_name, "Category & Documents provided", "Bated labels"]
    row_idx = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border_thin

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 30

    ws.freeze_panes = "A4"

    for need in [name_col_name, cat_col_name]:
        if need not in df.columns:
            df[need] = ""
    if bates_col_name not in df.columns:
        df[bates_col_name] = ""
    if date_col_name not in df.columns:
        df[date_col_name] = datetime.today().date()

    sdf = df.copy()
    if cat_col_name in sdf.columns:
        sdf[cat_col_name] = sdf[cat_col_name].fillna("")
        sdf.sort_values([cat_col_name, name_col_name], inplace=True, kind="stable")

    row = 4
    for cat, block in sdf.groupby(cat_col_name, dropna=False):
        cat_text = str(cat) if str(cat).strip() else "Uncategorized"
        ws.cell(row=row, column=2, value=cat_text).fill = category_fill
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = left
        for c in (1, 2, 3):
            cc = ws.cell(row=row, column=c)
            cc.border = border_thin
        row += 1

        for _, r in block.iterrows():
            ws.cell(row=row, column=1, value=r.get(date_col_name, "")).alignment = center
            ws.cell(row=row, column=2, value=r.get(name_col_name, "")).alignment = left
            ws.cell(row=row, column=3, value=r.get(bates_col_name, "")).alignment = left
            for c in (1, 2, 3):
                ws.cell(row=row, column=c).font = normal_font
                ws.cell(row=row, column=c).border = border_thin
            row += 1

        row += 1

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    logger.info("build_discovery_xlsx: spreadsheet built — %d data row(s) written", row - 4)
    return out.getvalue()
